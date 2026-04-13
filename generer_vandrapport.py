"""
Generer vandbesparings-rapport (PDF) for en liste af norske sites.

Læser antalvask.xlsx (kolonner: "Site ", "Antal vask "),
slår kommune op via Kartverket (https://ws.geonorge.no/),
beregner besparelse pr. site og samlet, og genererer en PDF.

Brug:
    .venv/bin/python generer_vandrapport.py
        -> skriver vandrapport.pdf

Justér konstanterne nedenfor efter behov.
"""

import os
import re
import sys
import time
from urllib.parse import quote

import openpyxl
import urllib.request
import json
from datetime import datetime
from fpdf import FPDF


# ────────── KONFIGURATION ──────────
INPUT_FIL       = 'antalvask.xlsx'
OUTPUT_PDF      = 'vandrapport.pdf'

GENBRUG_PCT     = 80      # %
LITER_PR_VASK   = 400     # liter
SYSTEM_PRIS_NOK = 550000  # NOK pr. anlæg

# Vannpriser (NOK/m³ inkl. 15% mva) – speilet fra vandberegner.html, april 2026
VANDPRISER_NO = {
    "Oslo": 83, "Bergen": 58, "Trondheim": 49, "Stavanger": 26,
    "Kristiansand": 65, "Tromsø": 76, "Drammen": 93, "Fredrikstad": 72,
    "Sarpsborg": 68, "Sandnes": 32, "Bodø": 80, "Ålesund": 62,
    "Skien": 72, "Sandefjord": 68, "Tønsberg": 66, "Moss": 74,
    "Haugesund": 60, "Molde": 68, "Lillehammer": 70, "Gjøvik": 65,
    "Hamar": 66, "Arendal": 72, "Grimstad": 70, "Larvik": 66,
    "Horten": 68, "Halden": 82, "Steinkjer": 72, "Namsos": 78,
    "Alta": 84, "Harstad": 80, "Narvik": 74, "Rana": 74,
    "Kongsberg": 70, "Notodden": 74, "Porsgrunn": 76, "Bærum": 65,
    "Asker": 70, "Lillestrøm": 137, "Nordre Follo": 70, "Nesodden": 72,
    "Frogn": 74, "Ås": 68, "Nittedal": 70, "Ullensaker": 64,
    "Eidsvoll": 68, "Nannestad": 70, "Elverum": 72, "Ringsaker": 68,
    "Stange": 70, "Ringerike": 66, "Lørenskog": 78, "Vestby": 68,
}

# Manuel mapping for sites hvor stedsnavn ikke kan slås op via API
# (UnoX-stationer har ofte interne navne der ikke matcher offentlige stedsnavne)
MANUEL_KOMMUNE = {
    # Eksempler – udfyld efter behov når API returnerer 'Ukendt':
    # "Tempokrysset": "Trondheim",
    # "Drotningsvik": "Bergen",
}

DEFAULT_KOMMUNE_PRIS = 70  # NOK/m³ – brugt hvis kommune ikke findes i prislisten


# ────────── KARTVERKET API ──────────
def slaa_kommune_op(stedsnavn: str) -> str | None:
    """
    Slår et stedsnavn op via Kartverkets stedsnavn-API.
    Returnerer kommunenavnet eller None.
    """
    url = (
        f"https://ws.geonorge.no/stedsnavn/v1/navn"
        f"?sok={quote(stedsnavn)}&treffPerSide=5&fuzzy=true"
    )
    try:
        with urllib.request.urlopen(url, timeout=8) as resp:
            data = json.loads(resp.read())
        navne = data.get('navn', [])
        for n in navne:
            kommuner = n.get('kommuner', [])
            if kommuner:
                return kommuner[0].get('kommunenavn')
    except Exception as e:
        print(f"  ! API-fejl for '{stedsnavn}': {e}", file=sys.stderr)
    return None


def fald_tilbage_til_adresse(stedsnavn: str) -> str | None:
    """Fallback: prøv adresser/sok hvis stedsnavn-API ikke gav svar."""
    url = (
        f"https://ws.geonorge.no/adresser/v1/sok"
        f"?sok={quote(stedsnavn)}&treffPerSide=3&fuzzy=true"
    )
    try:
        with urllib.request.urlopen(url, timeout=8) as resp:
            data = json.loads(resp.read())
        adr = data.get('adresser', [])
        if adr:
            return adr[0].get('kommunenavn')
    except Exception:
        pass
    return None


# ────────── EXCEL ──────────
SITE_REGEX = re.compile(r'^\s*\(?(\d+)\)?\s+(.+?)\s*$')

def parse_site(raw: str):
    """Returnerer (site_nr, navn) — fjerner whitespace + parantes om nummer."""
    m = SITE_REGEX.match(raw)
    if m:
        return m.group(1), m.group(2).strip()
    return None, raw.strip()


def laes_excel(path: str):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # header
            continue
        if not row or row[0] is None or row[1] is None:
            continue
        site_nr, navn = parse_site(str(row[0]))
        try:
            antal = int(row[1])
        except (TypeError, ValueError):
            continue
        rows.append({'nr': site_nr, 'navn': navn, 'antal_vask': antal})
    return rows


# ────────── BEREGNING ──────────
def beregn_site(site, kommune, pris):
    total_m3      = site['antal_vask'] * LITER_PR_VASK / 1000
    sparet_m3     = total_m3 * GENBRUG_PCT / 100
    besparelse    = sparet_m3 * pris
    payback_aar   = SYSTEM_PRIS_NOK / besparelse if besparelse > 0 else None
    return {
        **site,
        'kommune':    kommune,
        'pris_m3':    pris,
        'total_m3':   total_m3,
        'sparet_m3':  sparet_m3,
        'besparelse': besparelse,
        'payback':    payback_aar,
    }


# ────────── PDF ──────────
# UNO-X / 7-Eleven Norge brand-farver
UNOX_GUL    = (255, 213, 0)    # #FFD500 – signaturgul
UNOX_SORT   = (20, 20, 20)
SEVEN_GROEN = (0, 122, 51)     # 7-Eleven grøn
SEVEN_ROED  = (220, 36, 31)    # 7-Eleven rød
SEVEN_ORANGE= (245, 130, 32)   # 7-Eleven orange
TEKST_MOERK = (30, 30, 35)
TEKST_GRAA  = (110, 110, 115)
LINJE_LYS   = (230, 230, 235)


class Rapport(FPDF):
    """A4 liggende; header + footer på alle sider undtagen forsiden."""
    forsidemodus = False
    logo_path    = None

    def header(self):
        if self.forsidemodus:
            return
        # Slank topbjælke i UNO-X gult
        self.set_fill_color(*UNOX_GUL)
        self.rect(0, 0, 297, 8, style='F')
        self.set_fill_color(*UNOX_SORT)
        self.rect(0, 8, 297, 1.2, style='F')
        # Logo + titel
        if self.logo_path and os.path.exists(self.logo_path):
            self.image(self.logo_path, x=10, y=12, h=10)
        self.set_xy(0, 13)
        self.set_font('DejaVu', 'B', 10)
        self.set_text_color(*UNOX_SORT)
        self.cell(0, 6, 'VANDBESPARELSE · UNO-X / 7-ELEVEN NORGE',
                  align='R', new_x='LMARGIN', new_y='NEXT')
        self.set_xy(0, 19)
        self.set_font('DejaVu', '', 7.5)
        self.set_text_color(*TEKST_GRAA)
        self.cell(0, 4, 'Udarbejdet af AhSolution ApS', align='R')
        self.set_y(26)

    def footer(self):
        if self.forsidemodus:
            return
        self.set_y(-12)
        self.set_draw_color(*UNOX_GUL)
        self.set_line_width(0.6)
        self.line(10, self.get_y(), 287, self.get_y())
        self.ln(1)
        self.set_font('DejaVu', '', 7)
        self.set_text_color(*TEKST_GRAA)
        self.cell(95, 6, 'AhSolution ApS · CVR 45081125 · ah@ahsolution.dk', align='L')
        self.cell(97, 6, 'Vandbesparingsrapport · UNO-X / 7-Eleven Norge', align='C')
        self.cell(95, 6, f'Side {self.page_no()}', align='R')


def fmt_kr(n): return f"{n:,.0f}".replace(",", ".") + " kr"
def fmt_m3(n): return f"{n:,.0f}".replace(",", ".") + " m³"
def fmt_int(n): return f"{n:,.0f}".replace(",", ".")
def fmt_aar(n):
    if n is None: return "–"
    return f"{n:.1f} år"


def _tegne_noegletal_kort(pdf, x, y, w, h, label, value, baggrund, tekstfarve, accent=None):
    """Et flot 'KPI-kort' med farvet venstrekant."""
    pdf.set_fill_color(*baggrund)
    pdf.rect(x, y, w, h, style='F')
    if accent:
        pdf.set_fill_color(*accent)
        pdf.rect(x, y, 1.6, h, style='F')
    pdf.set_xy(x + 4, y + 2.5)
    pdf.set_font('DejaVu', '', 7.5)
    pdf.set_text_color(*TEKST_GRAA)
    pdf.cell(w - 6, 4, label.upper())
    pdf.set_xy(x + 4, y + 7)
    pdf.set_font('DejaVu', 'B', 14)
    pdf.set_text_color(*tekstfarve)
    pdf.cell(w - 6, 8, value)


def _tegne_forside(pdf, total_sites, total_antal, total_sparet_m3, total_besparelse, samlet_payback, samlet_investering):
    """Stor, branded forside med UNO-X gul header."""
    pdf.forsidemodus = True
    pdf.add_page()

    # Toppen: stor gul flade
    pdf.set_fill_color(*UNOX_GUL)
    pdf.rect(0, 0, 297, 95, style='F')
    pdf.set_fill_color(*UNOX_SORT)
    pdf.rect(0, 95, 297, 2, style='F')

    # Logo top venstre
    if pdf.logo_path and os.path.exists(pdf.logo_path):
        pdf.image(pdf.logo_path, x=18, y=14, h=14)

    # 7-Eleven farvebar (røde/grønne/orange striber) øverst til højre
    bar_x = 230; bar_y = 16
    for i, c in enumerate([SEVEN_ROED, SEVEN_GROEN, SEVEN_ORANGE]):
        pdf.set_fill_color(*c)
        pdf.rect(bar_x + i * 12, bar_y, 10, 5, style='F')
    pdf.set_xy(bar_x - 30, bar_y + 6)
    pdf.set_font('DejaVu', 'B', 7.5)
    pdf.set_text_color(*UNOX_SORT)
    pdf.cell(80, 4, 'UNO-X · 7-ELEVEN NORGE', align='R')

    # Stor titel
    pdf.set_xy(18, 42)
    pdf.set_font('DejaVu', 'B', 30)
    pdf.set_text_color(*UNOX_SORT)
    pdf.cell(0, 13, 'Vannbesparingsrapport', new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(18)
    pdf.set_font('DejaVu', '', 14)
    pdf.cell(0, 7, f'Estimeret potentiale på tværs af {total_sites} norske bilvask-sites')
    pdf.set_xy(18, 78)
    pdf.set_font('DejaVu', '', 9.5)
    pdf.set_text_color(60)
    pdf.cell(0, 5, f'Udarbejdet af AhSolution ApS · {datetime.now().strftime("%B %Y").capitalize()}')

    # Nøgletals-blok midt på siden
    pdf.set_y(115)
    pdf.set_x(18)
    pdf.set_font('DejaVu', 'B', 11)
    pdf.set_text_color(*UNOX_SORT)
    pdf.cell(0, 6, 'Sammenfatning – årligt potentiale', new_x='LMARGIN', new_y='NEXT')
    pdf.set_draw_color(*UNOX_GUL)
    pdf.set_line_width(1.2)
    pdf.line(18, pdf.get_y() + 1, 60, pdf.get_y() + 1)
    pdf.ln(7)

    # 4 store nøgletal-kort
    kort_w = 62; kort_h = 22; gap = 4
    y0 = pdf.get_y()
    cards = [
        ('Antal sites',           fmt_int(total_sites),      (250, 250, 252), UNOX_SORT,  UNOX_SORT),
        ('Antal vask pr. år',     fmt_int(total_antal),      (250, 250, 252), UNOX_SORT,  SEVEN_ORANGE),
        ('Vand sparet pr. år',    fmt_m3(total_sparet_m3),   (235, 247, 240), (0, 90, 40),  SEVEN_GROEN),
        ('Årlig besparelse',      fmt_kr(total_besparelse),  (255, 251, 222), UNOX_SORT,  UNOX_GUL),
    ]
    x0 = 18
    for i, (lbl, val, bg, tx, ac) in enumerate(cards):
        _tegne_noegletal_kort(pdf, x0 + i * (kort_w + gap), y0, kort_w, kort_h, lbl, val, bg, tx, ac)

    # Sekundære nøgletal nedenunder
    y1 = y0 + kort_h + gap
    sub_w = 84
    _tegne_noegletal_kort(pdf, 18,                  y1, sub_w, 18, 'Samlet investering', fmt_kr(samlet_investering), (250, 250, 252), UNOX_SORT, UNOX_SORT)
    _tegne_noegletal_kort(pdf, 18 + sub_w + gap,    y1, sub_w, 18, 'Samlet payback',     fmt_aar(samlet_payback),    (235, 247, 240), (0, 90, 40), SEVEN_GROEN)
    _tegne_noegletal_kort(pdf, 18 + 2*(sub_w+gap),  y1, sub_w, 18, 'Pris pr. anlæg',     fmt_kr(SYSTEM_PRIS_NOK),    (255, 251, 222), UNOX_SORT, UNOX_GUL)

    # Footer-band på forside
    pdf.set_y(-32)
    pdf.set_fill_color(*UNOX_SORT)
    pdf.rect(0, pdf.get_y(), 297, 32, style='F')
    pdf.set_xy(18, pdf.get_y() + 8)
    pdf.set_font('DejaVu', 'B', 9)
    pdf.set_text_color(*UNOX_GUL)
    pdf.cell(0, 5, 'AHSOLUTION APS', new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(18)
    pdf.set_font('DejaVu', '', 8.5)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 4.5, 'Idea · Solution · Execution', new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(18)
    pdf.set_font('DejaVu', '', 7.5)
    pdf.set_text_color(200)
    pdf.cell(0, 4, 'Tingbakken 39, 8883 Gjern · ah@ahsolution.dk · +45 23 81 72 72 · CVR 45081125')

    pdf.forsidemodus = False


def generer_pdf(resultater, output_path):
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    FONT_DIR = os.path.join(BASE_DIR, 'fonts')
    LOGO     = os.path.join(BASE_DIR, 'static', 'logo.png')

    pdf = Rapport(orientation='L', format='A4')
    pdf.logo_path = LOGO
    pdf.add_font('DejaVu', '',  os.path.join(FONT_DIR, 'DejaVuSans.ttf'))
    pdf.add_font('DejaVu', 'B', os.path.join(FONT_DIR, 'DejaVuSans-Bold.ttf'))
    pdf.add_font('DejaVu', 'I', os.path.join(FONT_DIR, 'DejaVuSans-Oblique.ttf'))
    pdf.set_auto_page_break(True, margin=18)

    # ── Aggregater ──
    total_sites       = len(resultater)
    total_antal       = sum(r['antal_vask'] for r in resultater)
    total_m3          = sum(r['total_m3']   for r in resultater)
    total_sparet_m3   = sum(r['sparet_m3']  for r in resultater)
    total_besparelse  = sum(r['besparelse'] for r in resultater)
    samlet_invest     = SYSTEM_PRIS_NOK * total_sites
    samlet_payback    = samlet_invest / total_besparelse if total_besparelse > 0 else None

    # ── Forside ──
    _tegne_forside(pdf, total_sites, total_antal, total_sparet_m3,
                   total_besparelse, samlet_payback, samlet_invest)

    # ── Side 2: Metode + forudsætninger ──
    pdf.add_page()
    pdf.set_font('DejaVu', 'B', 14)
    pdf.set_text_color(*UNOX_SORT)
    pdf.cell(0, 8, 'Om rapporten', new_x='LMARGIN', new_y='NEXT')
    pdf.set_draw_color(*UNOX_GUL); pdf.set_line_width(1.2)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.l_margin + 35, pdf.get_y())
    pdf.ln(4)

    pdf.set_font('DejaVu', '', 9.5)
    pdf.set_text_color(*TEKST_MOERK)
    pdf.multi_cell(0, 5,
        'Denne rapport estimerer den årlige vannbesparelse ved at installere et komplet '
        'vanngjenbrukssystem fra AhSolution på UNO-X / 7-Eleven Norges bilvask-sites. '
        'For hvert site er kommunen slået op via Kartverkets stedsnavn-API '
        '(ws.geonorge.no), og den variable vannpris pr. m³ (vann + avløp inkl. 15 % mva) '
        'er hentet fra de officielle kommunale gebyrregulativ for 2026.\n\n'
        'Bekræftede 2026-priser er anvendt for Oslo, Bergen, Trondheim, Stavanger, '
        'Drammen, Asker og Lillestrøm. For øvrige kommuner er der anvendt regionale '
        'estimater baseret på SSB-data og lokale gebyrregulativer. Sites hvor stedsnavnet '
        'ikke kunne mappes entydigt til en kommune, bruger en konservativ default-pris.\n\n'
        'Antal vask pr. site er trukket direkte fra UNO-X / 7-Elevens egne data '
        '(antalvask.xlsx).')
    pdf.ln(3)

    # Forudsætninger – som boks med gul venstrekant
    box_y = pdf.get_y()
    pdf.set_fill_color(252, 252, 245)
    pdf.rect(pdf.l_margin, box_y, 287 - pdf.l_margin, 36, style='F')
    pdf.set_fill_color(*UNOX_GUL)
    pdf.rect(pdf.l_margin, box_y, 2, 36, style='F')
    pdf.set_xy(pdf.l_margin + 6, box_y + 3)
    pdf.set_font('DejaVu', 'B', 10.5)
    pdf.set_text_color(*UNOX_SORT)
    pdf.cell(0, 5, 'Forudsætninger i beregningen', new_x='LMARGIN', new_y='NEXT')
    pdf.set_x(pdf.l_margin + 6)
    pdf.set_font('DejaVu', '', 9)
    pdf.set_text_color(*TEKST_MOERK)
    forud = [
        ('Vandgenbrug',          f'{GENBRUG_PCT} %',                    'andel af vand der renses og genbruges på sitet'),
        ('Vandforbrug pr. vask', f'{LITER_PR_VASK} liter',              'typisk for en standard bilvask'),
        ('Systempris',           fmt_kr(SYSTEM_PRIS_NOK),                'pr. site, inkl. installation'),
        ('Payback',              'Investering ÷ årlig besparelse',       'simpel tilbagebetalingstid'),
    ]
    for label, val, beskr in forud:
        pdf.set_x(pdf.l_margin + 8)
        pdf.set_font('DejaVu', 'B', 9)
        pdf.cell(48, 5.5, '• ' + label, align='L')
        pdf.set_font('DejaVu', '', 9)
        pdf.cell(45, 5.5, val, align='L')
        pdf.set_text_color(*TEKST_GRAA)
        pdf.cell(0, 5.5, beskr, new_x='LMARGIN', new_y='NEXT')
        pdf.set_text_color(*TEKST_MOERK)
    pdf.ln(8)

    # Bemærkning
    pdf.set_font('DejaVu', 'I', 8.5)
    pdf.set_text_color(*TEKST_GRAA)
    pdf.multi_cell(0, 4.4,
        'Bemærk: Beregningen er vejledende. Faktiske vannpriser kan variere fra estimaterne, og den reelle '
        'besparelse afhænger af installation, vandkvalitet og driftsforhold. Norske vann- og avløpsgebyrer '
        'er steget kraftigt de seneste år og forventes at stige yderligere frem mod 2028 — den faktiske '
        'besparelse vil derfor sandsynligvis vokse over tid. Kontakt AhSolution for et konkret tilbud '
        'og dimensionering af det enkelte site.')

    # ── Side 3+: Detaljer pr. site ──
    pdf.add_page()
    pdf.set_font('DejaVu', 'B', 14)
    pdf.set_text_color(*UNOX_SORT)
    pdf.cell(0, 8, 'Detaljer pr. site', new_x='LMARGIN', new_y='NEXT')
    pdf.set_draw_color(*UNOX_GUL); pdf.set_line_width(1.2)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.l_margin + 35, pdf.get_y())
    pdf.ln(2)
    pdf.set_font('DejaVu', '', 8.5)
    pdf.set_text_color(*TEKST_GRAA)
    pdf.cell(0, 5, f'{total_sites} sites · sorteret efter årlig besparelse', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(2)

    # Tabel-kolonner
    cols = [
        ('Nr.',         15,  'L'),
        ('Site',        58,  'L'),
        ('Kommune',     38,  'L'),
        ('Antal vask',  24,  'R'),
        ('kr/m³',       18,  'R'),
        ('Vand sparet', 28,  'R'),
        ('Besparelse',  32,  'R'),
        ('Payback',     22,  'R'),
    ]
    total_w = sum(w for _, w, _ in cols)

    def tabel_header():
        pdf.set_fill_color(*UNOX_SORT)
        pdf.set_text_color(*UNOX_GUL)
        pdf.set_font('DejaVu', 'B', 8.5)
        for label, w, align in cols:
            pdf.cell(w, 7.5, ' ' + label if align == 'L' else label + ' ', fill=True, align=align)
        pdf.ln()

    tabel_header()

    # Sortér efter besparelse faldende
    resultater.sort(key=lambda r: r['besparelse'], reverse=True)

    pdf.set_text_color(*TEKST_MOERK)
    pdf.set_font('DejaVu', '', 8.5)
    for i, r in enumerate(resultater):
        # Sidebreaks: gentag header
        if pdf.get_y() > 185:
            pdf.add_page()
            tabel_header()
            pdf.set_text_color(*TEKST_MOERK)
            pdf.set_font('DejaVu', '', 8.5)

        bg = (252, 252, 248) if i % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(*bg)
        kommune_vis = r['kommune'] or '⚠ Ukendt'
        cells_data = [
            r['nr'] or '',
            r['navn'][:32],
            kommune_vis,
            fmt_int(r['antal_vask']),
            f"{r['pris_m3']:.0f}",
            fmt_m3(r['sparet_m3']),
            fmt_kr(r['besparelse']),
            fmt_aar(r['payback']),
        ]
        for (label, w, align), txt in zip(cols, cells_data):
            txt_padded = ' ' + str(txt) if align == 'L' else str(txt) + ' '
            pdf.cell(w, 6.5, txt_padded, fill=True, align=align)
        pdf.ln()

    # Total-række nederst
    pdf.set_fill_color(*UNOX_GUL)
    pdf.set_text_color(*UNOX_SORT)
    pdf.set_font('DejaVu', 'B', 9)
    total_cells = ['', 'TOTAL', f'{total_sites} sites',
                   fmt_int(total_antal), '', fmt_m3(total_sparet_m3),
                   fmt_kr(total_besparelse), fmt_aar(samlet_payback)]
    for (label, w, align), txt in zip(cols, total_cells):
        txt_padded = ' ' + str(txt) if align == 'L' else str(txt) + ' '
        pdf.cell(w, 8, txt_padded, fill=True, align=align)
    pdf.ln()

    # Noter om manglende kommuner
    ukendte = [r for r in resultater if not r['kommune']]
    if ukendte:
        pdf.ln(5)
        pdf.set_font('DejaVu', 'I', 8)
        pdf.set_text_color(*SEVEN_ROED)
        pdf.multi_cell(0, 4.5,
            f'⚠ {len(ukendte)} site(s) kunne ikke knyttes til en kommune via Kartverket. '
            f'Default-pris ({DEFAULT_KOMMUNE_PRIS} kr/m³) er anvendt. '
            f'Tilføj dem til MANUEL_KOMMUNE i scriptet for præcis pris.')

    pdf.output(output_path)


# ────────── MAIN ──────────
def main():
    print(f"Læser {INPUT_FIL} …")
    sites = laes_excel(INPUT_FIL)
    print(f"  fundet {len(sites)} sites\n")

    print("Slår kommuner op via Kartverket …")
    resultater = []
    for s in sites:
        if s['navn'] in MANUEL_KOMMUNE:
            kommune = MANUEL_KOMMUNE[s['navn']]
            kilde = 'manuel'
        else:
            kommune = slaa_kommune_op(s['navn']) or fald_tilbage_til_adresse(s['navn'])
            kilde = 'API'
            time.sleep(0.15)  # vær snill mod API'et
        pris = VANDPRISER_NO.get(kommune, DEFAULT_KOMMUNE_PRIS) if kommune else DEFAULT_KOMMUNE_PRIS
        marker = '✓' if kommune else '✗'
        print(f"  {marker} ({s['nr']}) {s['navn']:<28} → {kommune or 'UKENDT':<20} ({pris} kr/m³, {kilde})")
        resultater.append(beregn_site(s, kommune, pris))

    print(f"\nGenererer {OUTPUT_PDF} …")
    generer_pdf(resultater, OUTPUT_PDF)
    print(f"  ✓ skrevet til {os.path.abspath(OUTPUT_PDF)}")


if __name__ == '__main__':
    main()
